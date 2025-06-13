#!/usr/bin/env python3

import os
import re
import openpyxl
import math
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from collections import defaultdict

# Aggregates runtime metrics for the MSA benchmark results

# Directory to search
#DATA_DIR = '../data/de-novo'
DATA_DIR = '../data/reference'

# Filename to save the runtime metrics to
#OUTPUT_FILE = 'centralized_de_novo_msa_runtime_metrics.xlsx'
OUTPUT_FILE = 'centralized_reference_msa_runtime_metrics.xlsx'

# Create a new workbook and add two worksheets
wb = openpyxl.Workbook()

# Rename the default sheet to 'stats'
stats_sheet = wb.active
stats_sheet.title = 'stats'

# Add the 'reads' worksheet
reads_sheet = wb.create_sheet(title='reads')

# Data to insert into 'reads' worksheet
reads_data = [
    ['0.1k', 91],
    ['0.5k', 455],
    ['1k', 898],
    ['10k', 8965],
    ['50k', 45059],
    ['100k', 90192],
    ['200k', 180206],
    ['500k', 449560],
    ['1m', 899192],
    ['2m', 1798684],
]

# Optionally add headers
reads_sheet.append(['Număr de citiri proiectate', 'Număr final de citiri'])

# Insert the data
for row in reads_data:
    reads_sheet.append(row)

# Patterns for log files
LOG_PATTERN = re.compile(r'benchmark_(\d+)_([a-z]+)\.log$')

# Data structure to hold parsed results
benchmark_data = {}

# Recursively walk through all subdirectories
for root, dirs, files in os.walk(DATA_DIR):
    for file in files:
        # Accession runs are only used for the accuracy evaluation. The accession mapping adds time to the execution.
        if 'kraken2-accession' in root:
            continue
        match = LOG_PATTERN.match(file)
        if match:
            number, logtype = match.groups()
            number = int(number)
            if number not in benchmark_data:
                benchmark_data[number] = {}
            benchmark_data[number][logtype] = os.path.join(root, file)

# Now, for each [number], parse the required files if present
def parse_detalii(path):
    cores = 32
    ram = 49152
    with open(path, encoding='utf-8') as f:
        for line in f:
            if 'Număr de nuclee' in line:
                try:
                    cores = int(re.search(r'(\d+)', line).group(1))
                except Exception:
                    pass
            elif 'Memorie RAM (MB)' in line:
                try:
                    ram = int(re.search(r'(\d+)', line).group(1))
                except Exception:
                    pass
    return {'cores': cores, 'RAM': ram}

def parse_disc(path):
    max_delta = None
    final_delta = None
    with open(path, encoding='utf-8') as f:
        for line in f:
            if 'DELTA_ROOT' in line and 'FINAL' not in line:
                try:
                    max_delta = int(re.search(r'(-?\d+)', line).group(1))
                except Exception:
                    pass
            elif 'DELTA_FINAL_ROOT' in line:
                try:
                    final_delta = int(re.search(r'(-?\d+)', line).group(1))
                except Exception:
                    pass
    return {'max delta': max_delta, 'final delta': final_delta}

def parse_timp(path):
    result = {}
    with open(path, encoding='utf-8') as f:
        first_line = f.readline().strip()
        # Check if command was terminated by signal
        if first_line.startswith("Command terminated by signal"):
            return None
        
        # Reset file pointer to beginning and continue processing
        f.seek(0)
        for line in f:
            if 'Command being timed:' in line:
                cmd = re.search(r'"(.+?)"', line)
                if cmd:
                    command = cmd.group(1)
                    parts = command.split()
                    result['command'] = command
                    result['executable'] = os.path.basename(parts[0]) if parts else ''
                    result['params'] = ' '.join(parts[1:]) if len(parts) > 1 else ''
                    if 'nr_' in result['params'].lower():
                        result['executable'] = result['executable'] + '-nr'
                    elif 'nr99_' in result['params'].lower():
                        result['executable'] = result['executable'] + '-nr99'
                    elif 'seed_' in result['params'].lower():
                        result['executable'] = result['executable'] + '-seed'
            elif 'User time (seconds):' in line:
                result['user time'] = float(line.split(':', 1)[1].strip())
            elif 'System time (seconds):' in line:
                result['system time'] = float(line.split(':', 1)[1].strip())
            elif 'Percent of CPU this job got:' in line:
                result['CPU usage'] = float(line.split(':', 1)[1].strip().replace('%',''))
            elif 'Elapsed (wall clock) time (h:mm:ss or m:ss)' in line:
                duration = line.split(':', 4)[4].strip()
                result['Duration'] = duration
            elif 'Maximum resident set size (kbytes):' in line:
                result['Max RAM'] = int(line.split(':', 1)[1].strip())
            elif 'File system inputs:' in line:
                result['File system inputs'] = int(line.split(':', 1)[1].strip())
            elif 'File system outputs:' in line:
                result['File system outputs'] = int(line.split(':', 1)[1].strip())
            elif 'Exit status:' in line:
                result['Exit status'] = int(line.split(':', 1)[1].strip())
    # Compute File system I/O
    if 'File system inputs' in result and 'File system outputs' in result:
        result['File system I/O'] = result['File system inputs'] + result['File system outputs']
    return result

# Aggregate all parsed data
parsed_benchmarks = []
for number, logs in benchmark_data.items():
    entry = {'number': number}
    if 'detalii' in logs:
        entry.update(parse_detalii(logs['detalii']))
    if 'disc' in logs:
        entry.update(parse_disc(logs['disc']))
    if 'timp' in logs:
        timp_result = parse_timp(logs['timp'])
        if timp_result is None:
            # Skip this entry if command was terminated by signal
            continue
        entry.update(timp_result)
    # Optionally, add more parsing for 'erori' and 'iesire' if needed
    parsed_benchmarks.append(entry)

# At this point, parsed_benchmarks contains all the aggregated data for further processing

# Helper: parse duration string to seconds
# Accepts formats like 'h:mm:ss', 'm:ss', or 's.ss'
def parse_duration_to_seconds(duration):
    if not duration:
        return None
    duration = duration.strip()
    if re.match(r'^\d+$', duration):
        return int(duration)
    parts = duration.split(':')
    try:
        if len(parts) == 3:
            h, m, s = parts
            return int(h) * 3600 + int(m) * 60 + float(s)
        elif len(parts) == 2:
            m, s = parts
            return int(m) * 60 + float(s)
        else:
            return float(duration)
    except Exception:
        return None

# Reads mapping from the reads worksheet
reads_mapping = {
    '0.1k': 91,
    '0.5k': 455,
    '1k': 898,
    '5k': 8965,
    '10k': 45059,
    '50k': 90192,
    '100k': 180206,
    '500k': 449560,
    '1m': 899192,
    '2m': 1798684,
}

# Extract unique executables
unique_executables = sorted(set(
    entry.get('executable', '') for entry in parsed_benchmarks if entry.get('executable', '')
))

# For each entry, compute Reads and Config
for entry in parsed_benchmarks:
    # Extract Reads from params using regex
    reads_match = re.search(r'([\d\.]+[km])\.fasta', entry.get('params', ''))
    reads = reads_match.group(1) if reads_match else ''
    entry['Reads'] = reads
    # Compute Config
    cores = entry.get('cores', 32)
    ram = entry.get('RAM', 49152)
    ram_gb = int(round(ram / 1024))
    # Fix RAM to 48GB if needed to account for variations in the VM RAM allocation
    if (ram_gb >= 47 and ram_gb <= 49):
        ram_gb = 48
    entry['Config'] = f"{reads}-{cores}CPU-{ram_gb}GB"

# Group entries by executable and then by Reads (for block layout)
exec_reads_blocks = defaultdict(lambda: defaultdict(list))
for entry in parsed_benchmarks:
    exe = entry.get('executable', '')
    reads = entry.get('Reads', '')
    if exe and reads:
        exec_reads_blocks[exe][reads].append(entry)

# Helper to find the row in the reads worksheet for a given Reads value
def get_reads_row(reads_value):
    for row in range(2, reads_sheet.max_row + 1):
        if reads_sheet.cell(row=row, column=1).value == reads_value:
            return row
    return None

# Helper to get sort key for Config
reads_order = list(reads_mapping.keys())
def config_sort_key(config):
    # config format: "{reads}-{cores}CPU-{ram_gb}GB"
    m = re.match(r'([\d\.]+[km])-(\d+)CPU-(\d+)GB', config)
    if m:
        reads, cpu, ram = m.groups()
        reads_idx = reads_order.index(reads) if reads in reads_order else math.inf
        return (-int(cpu), -int(ram), reads_idx)
    return (math.inf, math.inf, math.inf)

# Collect unique Config entries
configurations = sorted({entry['Config'] for entry in parsed_benchmarks if entry.get('Config')}, key=config_sort_key)

# Map (exe, config) to entry for fast lookup
exe_config_map = {}
for entry in parsed_benchmarks:
    exe = entry.get('executable', '')
    config = entry.get('Config', '')
    if exe and config:
        exe_config_map[(exe, config)] = entry

# Prepare to write the stats worksheet
# Clear the stats worksheet
for row in stats_sheet.iter_rows(min_row=1, max_row=stats_sheet.max_row):
    for cell in row:
        cell.value = None

col = 2  # Start from column B, column A is for executable
header_rows = 2

# Write config headers and block headers
for config in configurations:
    stats_sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+5)
    stats_sheet.cell(row=1, column=col, value=config)
    stats_sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
    headers = ['Duration (s)', 'Speed (seq/min)', 'CPU usage (%)', 'Max RAM (kB)', 'IO ops', 'Disk usage (final delta)']
    for i, h in enumerate(headers):
        stats_sheet.cell(row=2, column=col+i, value=h)
    col += 6

# Write data rows
row_idx = 3
for exe in unique_executables:
    stats_sheet.cell(row=row_idx, column=1, value=exe)
    col = 2
    for config in configurations:
        entry = exe_config_map.get((exe, config))
        # Only add data if exit status is zero or not present
        if entry and (entry.get('Exit status', 0) == 0):
            duration_sec = parse_duration_to_seconds(entry.get('Duration', ''))
            stats_sheet.cell(row=row_idx, column=col, value=duration_sec)
            reads = entry.get('Reads', '')
            reads_row = get_reads_row(reads)
            duration_cell = f"{get_column_letter(col)}{row_idx}"
            reads_cell = f"reads!$B${reads_row}" if reads_row else None
            if reads_cell and duration_sec:
                speed_formula = f"=({reads_cell}*60)/{duration_cell}"
                stats_sheet.cell(row=row_idx, column=col+1, value=speed_formula)
            stats_sheet.cell(row=row_idx, column=col+2, value=entry.get('CPU usage'))
            stats_sheet.cell(row=row_idx, column=col+3, value=entry.get('Max RAM'))
            stats_sheet.cell(row=row_idx, column=col+4, value=entry.get('File system I/O'))
            stats_sheet.cell(row=row_idx, column=col+5, value=entry.get('final delta'))
        col += 6
    row_idx += 1

stats_sheet.freeze_panes = 'B3'

wb.save(OUTPUT_FILE)
